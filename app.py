from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
from datetime import datetime, timedelta
import random
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tempfile
import math

app = Flask(__name__, static_folder='static')

# Create a directory for storing shifts if it doesn't exist
SHIFTS_DIR = 'shifts'
if not os.path.exists(SHIFTS_DIR):
    os.makedirs(SHIFTS_DIR)

def save_shift(shift_data, start_date):
    """Save shift data to a JSON file"""
    filename = f"shift_{start_date}.json"
    filepath = os.path.join(SHIFTS_DIR, filename)
    
    # Check if a shift already exists for this week
    if os.path.exists(filepath):
        return False, "شیفت برای این هفته قبلاً تعریف شده است"
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(shift_data, f, ensure_ascii=False)
    return True, "شیفت با موفقیت ذخیره شد"

def get_saved_shifts():
    """Get list of all saved shifts"""
    shifts = []
    for filename in os.listdir(SHIFTS_DIR):
        if filename.endswith('.json'):
            filepath = os.path.join(SHIFTS_DIR, filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                shift_data = json.load(f)
                shifts.append({
                    'filename': filename,
                    'start_date': filename.replace('shift_', '').replace('.json', ''),
                    'data': shift_data
                })
    return shifts

def delete_shift(start_date):
    """Delete a saved shift"""
    filename = f"shift_{start_date}.json"
    filepath = os.path.join(SHIFTS_DIR, filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        return True, "شیفت با موفقیت حذف شد"
    return False, "شیفت مورد نظر یافت نشد"

def update_shift(start_date, shift_data):
    """Update an existing shift"""
    filename = f"shift_{start_date}.json"
    filepath = os.path.join(SHIFTS_DIR, filename)
    if not os.path.exists(filepath):
        return False, "شیفت مورد نظر یافت نشد"
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(shift_data, f, ensure_ascii=False)
    return True, "شیفت با موفقیت بروزرسانی شد"

def calculate_min_shifts_per_specialist(total_shifts, num_specialists, start_date, end_date):
    """Calculate minimum shifts per specialist on a monthly basis"""
    # تبدیل تاریخ‌ها به datetime
    start_dt = pd.to_datetime(start_date)
    end_dt = pd.to_datetime(end_date)
    
    # محاسبه تعداد روزهای هفته در این ماه
    days_in_current_week = (end_dt - start_dt).days + 1
    days_in_month = pd.Period(start_dt, freq='M').days_in_month
    
    # محاسبه نسبت روزهای این هفته به کل روزهای ماه
    week_to_month_ratio = days_in_current_week / days_in_month
    
    # محاسبه میانگین شیفت برای هر کارشناس در ماه
    shifts_per_specialist = total_shifts / num_specialists
    
    # اگر تعداد کارشناسان از نصف تعداد شیفت‌ها بیشتر باشد، نیازی به حداقل نیست
    if num_specialists > (total_shifts / 2):
        return 0
    
    # محاسبه حداقل شیفت برای این هفته بر اساس نسبت به ماه
    min_monthly_shifts = max(1, shifts_per_specialist / 4)  # حداقل یک شیفت در ماه
    min_shifts_this_week = round(min_monthly_shifts * week_to_month_ratio)
    
    return min_shifts_this_week

def convert_time_to_minutes(time_str):
    """Convert time string to minutes since midnight"""
    if pd.isna(time_str):
        return None
        
    try:
        # Try parsing as datetime object (Excel time format)
        if isinstance(time_str, datetime):
            return time_str.hour * 60 + time_str.minute
            
        # Try parsing as string in various formats
        for fmt in ['%I:%M:%S %p', '%H:%M:%S', '%H:%M']:
            try:
                time_obj = datetime.strptime(str(time_str), fmt)
                return time_obj.hour * 60 + time_obj.minute
            except ValueError:
                continue
                
        return None
    except Exception:
        return None

def is_specialist_available(specialist, shift_time, date, df):
    """Check if specialist is available for the given shift time on the given date"""
    # Convert shift time to minutes
    shift_hour, shift_minute = map(int, shift_time.split(':'))
    shift_time_minutes = shift_hour * 60 + shift_minute
    
    # Find specialist's schedule for this date
    specialist_schedule = df[
        (df['نام کارشناس'] == specialist) & 
        (df['تاریخ'].dt.date == date.date())
    ]
    
    if specialist_schedule.empty:
        return False
        
    row = specialist_schedule.iloc[0]
    
    # Get start and end times
    start_time = convert_time_to_minutes(row['زمان شروع'])
    end_time = convert_time_to_minutes(row['زمان پایان'])
    
    if start_time is None or end_time is None:
        print(f"Warning: Could not parse time for {specialist} on {date.date()}: Start={row['زمان شروع']}, End={row['زمان پایان']}")
        return False
    
    # Check if shift time falls within specialist's working hours
    # Also check if the entire 4-hour shift falls within working hours
    shift_end_minutes = shift_time_minutes + 240  # 4 hours in minutes
    
    is_available = start_time <= shift_time_minutes and shift_end_minutes <= end_time
    if not is_available:
        print(f"Debug: {specialist} not available at {shift_time} on {date.date()}")
        print(f"Work hours: {start_time}-{end_time} minutes, Shift: {shift_time_minutes}-{shift_end_minutes} minutes")
    
    return is_available

def load_holidays():
    try:
        with open('holidays.json', 'r', encoding='utf-8') as f:
            return json.load(f)['holidays']
    except FileNotFoundError:
        with open('holidays.json', 'w', encoding='utf-8') as f:
            json.dump({'holidays': []}, f, ensure_ascii=False, indent=4)
        return []

def save_holidays(holidays):
    with open('holidays.json', 'w', encoding='utf-8') as f:
        json.dump({'holidays': holidays}, f, ensure_ascii=False, indent=4)

@app.route('/api/holidays', methods=['GET'])
def get_holidays():
    return jsonify({'holidays': load_holidays()})

@app.route('/api/holidays', methods=['POST'])
def update_holidays():
    try:
        holidays = request.json.get('holidays', [])
        # اعتبارسنجی تاریخ‌ها
        for date in holidays:
            try:
                datetime.strptime(date, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': f'فرمت تاریخ {date} نامعتبر است'}), 400
        
        save_holidays(holidays)
        return jsonify({'success': True, 'message': 'روزهای تعطیل با موفقیت ذخیره شد'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def process_shifts(df, start_date, end_date):
    # تبدیل تاریخ‌ها به datetime
    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    
    # خواندن تنظیمات شیفت‌ها و روزهای تعطیل
    config = load_config()
    shifts = config['shifts']
    holidays = load_holidays()
    
    # دریافت لیست کارشناسان
    specialists = sorted(df['نام کارشناس'].unique().tolist())
    if not specialists:
        return {'error': 'هیچ کارشناسی در فایل اکسل یافت نشد'}
    
    # ساختار برای نگهداری برنامه و آمار
    schedule = {}
    specialist_shifts = {spec: [] for spec in specialists}  # لیست شیفت‌های هر کارشناس
    
    # آمار کلی
    total_shifts = 0  # تعداد کل شیفت‌ها
    holiday_count = 0  # تعداد روزهای تعطیل
    
    # ایجاد لیست تمام شیفت‌های ممکن
    all_shifts = []
    current = start
    while current <= end:
        date_str = current.strftime('%Y-%m-%d')
        
        if date_str in holidays:
            holiday_count += 1
        else:
            for shift in shifts:
                all_shifts.append({
                    'date': date_str,
                    'time': shift['time'],
                    'count': shift['count']
                })
                total_shifts += shift['count']  # تعداد کل شیفت‌های مورد نیاز
        
        # در هر صورت روز را در جدول نهایی اضافه کن
        schedule[date_str] = {shift['time']: [] for shift in shifts}
        current += timedelta(days=1)
    
    # برای هر کارشناس، شیفت‌های ممکن را پیدا کن
    specialist_available_shifts = {}
    for spec in specialists:
        available_shifts = []
        for shift in all_shifts:
            if is_specialist_available(spec, shift['time'], datetime.strptime(shift['date'], '%Y-%m-%d'), df):
                available_shifts.append(shift)
        specialist_available_shifts[spec] = available_shifts
    
    # تخصیص شیفت‌ها
    # مرحله 1: اطمینان از اینکه هر کارشناس حداقل یک شیفت دارد (اگر در دسترس باشد)
    random.shuffle(specialists)  # ترتیب تصادفی کارشناسان
    for spec in specialists:
        if specialist_available_shifts[spec]:  # اگر کارشناس شیفت در دسترس دارد
            # انتخاب تصادفی از بین شیفت‌های مناسب
            available_shifts = [
                shift for shift in specialist_available_shifts[spec]
                if len(schedule[shift['date']][shift['time']]) < shift['count']
            ]
            if available_shifts:
                shift = random.choice(available_shifts)
                schedule[shift['date']][shift['time']].append(spec)
                specialist_shifts[spec].append(f"{shift['date']} {shift['time']}")
    
    # مرحله 2: پر کردن بقیه شیفت‌ها به صورت عادلانه
    random.shuffle(all_shifts)  # ترتیب تصادفی شیفت‌ها
    for shift in all_shifts:
        date = shift['date']
        time = shift['time']
        needed = shift['count'] - len(schedule[date][time])
        
        if needed > 0:  # اگر هنوز به کارشناس نیاز داریم
            # کارشناسان در دسترس را بر اساس تعداد شیفت‌هایشان مرتب کن
            available_specs = [
                spec for spec in specialists 
                if is_specialist_available(spec, time, datetime.strptime(date, '%Y-%m-%d'), df) and
                f"{date} {time}" not in specialist_shifts[spec]
            ]
            
            # مرتب‌سازی بر اساس تعداد شیفت و کمی رندوم
            available_specs.sort(key=lambda x: len(specialist_shifts[x]) + random.random() * 0.5)
            
            # تخصیص شیفت به کارشناسان
            for i in range(min(needed, len(available_specs))):
                spec = available_specs[i]
                schedule[date][time].append(spec)
                specialist_shifts[spec].append(f"{date} {time}")
    
    # محاسبه آمار
    used_specialists = sum(1 for spec in specialists if specialist_shifts[spec])  # تعداد کارشناسان استفاده شده
    unused_specialists = len(specialists) - used_specialists  # تعداد کارشناسان استفاده نشده
    filled_shifts = sum(  # تعداد شیفت‌های پر شده
        len(schedule[date][time])
        for date in schedule
        for time in schedule[date]
    )
    
    # اضافه کردن آمار به خروجی
    stats = {
        'total_specialists': len(specialists),
        'used_specialists': used_specialists,
        'unused_specialists': unused_specialists,
        'total_shifts': total_shifts,
        'filled_shifts': filled_shifts,
        'holiday_count': holiday_count
    }
    
    return {'schedule': schedule, 'stats': stats}

def load_config():
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        # تنظیمات پیش‌فرض
        default_config = {
            "shifts": [
                {"time": "08:00", "count": 3},
                {"time": "08:30", "count": 3},
                {"time": "09:00", "count": 1},
                {"time": "12:00", "count": 1},
                {"time": "12:30", "count": 2},
                {"time": "13:00", "count": 3}
            ]
        }
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(default_config, f, ensure_ascii=False, indent=4)
        return default_config

def save_config(config):
    with open('config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'})
    
    file = request.files['file']
    start_date = request.form['start_date']
    end_date = request.form['end_date']
    
    try:
        # Read Excel file
        df = pd.read_excel(file)
        
        # Process the schedule
        result = process_shifts(df, start_date, end_date)
        
        # چک کردن وجود شیفت قبلی برای بازیابی آمار
        filename = os.path.join(SHIFTS_DIR, f'{start_date}.json')
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as f:
                saved_data = json.load(f)
                if 'stats' in saved_data:
                    # استفاده از آمار ذخیره شده به جای محاسبه مجدد
                    result['stats'] = saved_data['stats']
        
        return jsonify(result)
    except Exception as e:
        print(f"Error processing schedule: {str(e)}")
        return jsonify({'error': 'خطا در پردازش فایل یا ایجاد برنامه'})

@app.route('/save_shift', methods=['POST'])
def save_shift_route():
    try:
        data = request.get_json()
        start_date = data['start_date']
        schedule = data['schedule']
        stats = data.get('stats', {})  # دریافت آمار از درخواست
        
        # چک کردن وجود شیفت قبلی
        filename = os.path.join(SHIFTS_DIR, f'{start_date}.json')
        if os.path.exists(filename):
            # خواندن شیفت قبلی
            with open(filename, 'r', encoding='utf-8') as f:
                old_shift = json.load(f)
            
            # مقایسه تاریخ‌ها
            if old_shift['start_date'] == start_date:
                # آپدیت شیفت موجود
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump({
                        'start_date': start_date,
                        'data': schedule,
                        'stats': stats  # ذخیره آمار
                    }, f, ensure_ascii=False, indent=2)
                return jsonify({'success': True, 'message': 'شیفت با موفقیت بروزرسانی شد'})
        
        # ذخیره شیفت جدید
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump({
                'start_date': start_date,
                'data': schedule,
                'stats': stats  # ذخیره آمار
            }, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': 'شیفت با موفقیت ذخیره شد'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'خطا در ذخیره شیفت: {str(e)}'})

@app.route('/get_shifts')
def get_saved_shifts():
    shifts = []
    try:
        for filename in os.listdir(SHIFTS_DIR):
            if filename.endswith('.json'):
                with open(os.path.join(SHIFTS_DIR, filename), 'r', encoding='utf-8') as f:
                    shift_data = json.load(f)
                    shifts.append(shift_data)
        return jsonify(shifts)
    except Exception as e:
        print(f"Error in get_saved_shifts: {str(e)}")  # لاگ برای دیباگ
        return jsonify([])

@app.route('/delete_shift/<start_date>', methods=['DELETE'])
def delete_shift_route(start_date):
    try:
        filename = os.path.join(SHIFTS_DIR, f'{start_date}.json')
        if os.path.exists(filename):
            os.remove(filename)
            return jsonify({'success': True, 'message': 'شیفت با موفقیت حذف شد'})
        return jsonify({'success': False, 'message': 'شیفت مورد نظر یافت نشد'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'خطا در حذف شیفت: {str(e)}'})

@app.route('/update_shift', methods=['POST'])
def update_shift_route():
    try:
        data = request.get_json()
        start_date = data['start_date']
        schedule = data['schedule']
        
        # ذخیره در فایل JSON
        filename = os.path.join(SHIFTS_DIR, f'{start_date}.json')
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump({
                'start_date': start_date,
                'data': schedule
            }, f, ensure_ascii=False, indent=2)
        
        return jsonify({'success': True, 'message': 'شیفت با موفقیت بروزرسانی شد'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'خطا در بروزرسانی شیفت: {str(e)}'})

@app.route('/check_shift_exists/<start_date>')
def check_shift_exists(start_date):
    try:
        filename = os.path.join(SHIFTS_DIR, f'{start_date}.json')
        exists = os.path.exists(filename)
        return jsonify({'exists': exists})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/get_shift/<start_date>')
def get_shift(start_date):
    try:
        filename = os.path.join(SHIFTS_DIR, f'{start_date}.json')
        if not os.path.exists(filename):
            return jsonify({'success': False, 'message': 'شیفت مورد نظر یافت نشد'})
        
        with open(filename, 'r', encoding='utf-8') as f:
            shift_data = json.load(f)
            return jsonify({
                'success': True,
                'data': shift_data['data'],
                'stats': shift_data.get('stats', {})  # بازگرداندن آمار ذخیره شده
            })
    except Exception as e:
        return jsonify({'success': False, 'message': f'خطا در بازیابی شیفت: {str(e)}'})

@app.route('/export_excel/<start_date>')
def export_excel(start_date):
    try:
        shifts_file = os.path.join(SHIFTS_DIR, f'{start_date}.json')
        print(f"Looking for file: {shifts_file}")
        if not os.path.exists(shifts_file):
            print(f"File not found at: {shifts_file}")
            return jsonify({'error': 'شیفت مورد نظر یافت نشد'})

        with open(shifts_file, 'r', encoding='utf-8') as f:
            shift_data = json.load(f)

        # ایجاد فایل اکسل
        wb = Workbook()
        ws = wb.active
        ws.title = "برنامه شیفت"

        # تنظیم استایل‌ها
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name='IRANSansMonoSpaced')
        cell_font = Font(name='IRANSansMonoSpaced')
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ستون‌های جدول
        headers = ['ساعت', 'شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه']
        shift_times = ["08:00", "08:30", "09:00", "12:00", "12:30", "13:00"]

        # نوشتن هدرها
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
            cell.border = border
            ws.column_dimensions[chr(64 + col)].width = 25

        # نوشتن داده‌ها
        dates = sorted(shift_data['data'].keys())
        for row, time in enumerate(shift_times, 2):
            ws.cell(row=row, column=1, value=time).font = cell_font
            ws.cell(row=row, column=1).alignment = align
            ws.cell(row=row, column=1).border = border
            
            for col, date in enumerate(dates, 2):
                cell = ws.cell(row=row, column=col)
                specialists = shift_data['data'][date].get(time, [])
                cell.value = ', '.join(specialists) if specialists else ''
                cell.font = cell_font
                cell.alignment = align
                cell.border = border

        # تنظیم ارتفاع سطرها
        for row in range(1, len(shift_times) + 2):
            ws.row_dimensions[row].height = 30

        # ذخیره فایل
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'shift_schedule_{start_date}.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/settings')
def settings():
    config = load_config()
    return render_template('settings.html', shifts=config['shifts'])

@app.route('/api/settings', methods=['GET'])
def get_settings():
    config = load_config()
    return jsonify(config)

@app.route('/api/settings', methods=['POST'])
def update_settings():
    try:
        new_config = request.get_json()
        # اعتبارسنجی داده‌ها
        if 'shifts' not in new_config:
            return jsonify({'error': 'تنظیمات نامعتبر'}), 400
        
        for shift in new_config['shifts']:
            if 'time' not in shift or 'count' not in shift:
                return jsonify({'error': 'فرمت شیفت نامعتبر'}), 400
            try:
                count = int(shift['count'])
                if count < 0:
                    return jsonify({'error': 'تعداد شیفت باید مثبت باشد'}), 400
                shift['count'] = count
            except ValueError:
                return jsonify({'error': 'تعداد شیفت باید عدد باشد'}), 400
        
        save_config(new_config)
        return jsonify({'success': True, 'message': 'تنظیمات با موفقیت ذخیره شد'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/shifts', methods=['POST'])
def create_shifts():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'فایل اکسل یافت نشد'})
        
        file = request.files['file']
        start_date = request.form.get('start_date')
        
        if not file or not start_date:
            return jsonify({'error': 'اطلاعات ناقص است'})
        
        # خواندن فایل اکسل
        df = pd.read_excel(file)
        
        # محاسبه تاریخ پایان (5 روز بعد از تاریخ شروع)
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = start + timedelta(days=4)
        end_date = end.strftime('%Y-%m-%d')
        
        # پردازش شیفت‌ها
        result = process_shifts(df, start_date, end_date)
        
        if isinstance(result, dict) and 'error' in result:
            return jsonify(result)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error in create_shifts: {str(e)}")
        return jsonify({'error': str(e)})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)
